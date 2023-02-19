import difflib
import random
from traceback import format_exc

import openpyxl
import requests
from bs4 import BeautifulSoup
from transliterate import translit

import operator
proxies = [
	''
]

def string_comparison(first_string, last_string):
	return difflib.SequenceMatcher(a=first_string, b=last_string).ratio()

def string_translit(strig):
	return translit(strig, language_code='ru', reversed=True).lower()

def dict_number_similarities_update(value, using_dict, using_data):
	if value not in using_dict[using_data].keys():
		using_dict[using_data].update({value:0})
	else:
		using_dict[using_data].update({value:using_dict[using_data][value]+1})

	return using_dict

def get_session():
    # создать HTTP‑сеанс
    session = requests.Session()
    # выбираем один случайный прокси
    proxy = random.choice(proxies)
    session.proxies = {"http": proxy, "https": proxy}
    return session, proxy

def get_product(url):
	try:
		request, proxy = get_session() # для использования прокси
		data = requests.get(url).text
		return data
	except requests.exceptions.ProxyError:
		print(f'Прокси - {proxy} не работает')
		return get_product()

	except Exception as e:
		print('Ошибка:\n', format_exc())

def get_data(url):
	try:
		data = get_product(url)
		if data != None:
			soup = BeautifulSoup(data, features="html.parser")
			container = soup.select('ul.breadcrumbs__list a.breadcrumbs__link')
			categories_data = []
			for link in container:
				category_text = link.get_text(strip=True)
				category_link = link.get_attribute_list('href')[0].split('/')[-1]
				categories_data.append([category_text, category_link])

			for category in categories_data:
				if category[0] == 'Главная':
					categories_data.remove(category)
			product_name = soup.select('h1.same-part-kt__header span')[1].get_text(strip=True)

			try:
				product_data = parce(categories_data[0][0], categories_data[0][1], categories_data[1][0], categories_data[1][1], product_name)
				return product_data
			except IndexError:
				return None
			

	except Exception as e:
		print('Ошибка:\n', format_exc())

def parce(main_category, main_category_link, last_category, last_category_link, product_name):
	path_to_file = 'commission_and_logistics.xlsx'
	search_text = main_category.lower() # Текст первый ссылки(категории)
	seacrh_link = main_category_link.lower() # Ссылка первой категории
	two_search_text = last_category.lower() # Текст второй ссылки(категории)
	two_seacrh_link = last_category_link.lower() # Ссылка второй категории
	product = product_name.lower() # Название продукта

	#print('Ищем:', search_text, seacrh_link, two_search_text, two_seacrh_link, product)

	wb = openpyxl.load_workbook(path_to_file)  # Грузим наш прайс-лист
	sheet_active = wb[wb.sheetnames[0]]  # Получаем список всех листов в файле и Начинаем работать с самым первым
	row_max = sheet_active.max_row  # Получаем количество столбцов
	column_max = sheet_active.max_column  # Получаем количество строк
	row_min = 1 # Переменная, отвечающая за номер строки
	column_min = 1 # Переменная, отвечающая за номер столбца

	number_similarities = {
		'commission_percentage':{},
		'cost_mono_and_mix_logistics':{},
		'cost_monopallet_logistics':{},
		'cost_storing_mono_and_mixed':{},
		'storage_cost_monopallets':{},
		'calculation_according_actual_dimensions_goods':{}
	}
	while row_min <= row_max:
		row_min_min = row_min
		row_min_min = str(row_min_min)

		row_data = sheet_active[row_min_min]

		category = row_data[0].value.lower()
		comparison_first_category = string_comparison(search_text, category)
		comparison_first_category_link = string_comparison(seacrh_link, string_translit(category))
		# Сравниваем категорию и равниваем ссылку категории
		if comparison_first_category > 0.6 and comparison_first_category_link > 0.6:
			
			number_similarities = dict_number_similarities_update(row_data[2].value, number_similarities, 'commission_percentage')
					
			""" # Сравниваем вторую категорию, ссылку второй категории и/или товар
			two_category = row_data[1].value.lower()
				
			result = string_comparison(two_category, two_search_text) # Текст второй категории с текстом второй категории товара

			result_two = string_comparison(two_seacrh_link, string_translit(two_category)) # Ссылку второй категории с преобразованной ссылкой второй категории

			if result > 0.42 or result_two > 0.42:
				number_similarities = dict_number_similarities_update(row_data[3].value, number_similarities, 'cost_mono_and_mix_logistics')

				number_similarities = dict_number_similarities_update(row_data[4].value, number_similarities, 'cost_monopallet_logistics')

				number_similarities = dict_number_similarities_update(row_data[5].value, number_similarities, 'cost_storing_mono_and_mixed')

				number_similarities = dict_number_similarities_update(row_data[6].value, number_similarities, 'storage_cost_monopallets')

				number_similarities = dict_number_similarities_update(row_data[7].value, number_similarities, 'calculation_according_actual_dimensions_goods') """

				
				
		column_min += 1
		row_min += 1

	product_data = {}
	for key, value in number_similarities.items():
		try:
			product_data.update({key:max(number_similarities[key].items(), key=operator.itemgetter(1))[0]})
		except ValueError:
			continue

	return product_data